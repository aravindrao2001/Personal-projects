import hashlib
import os
from pathlib import Path
from typing import List, Dict

from bs4 import BeautifulSoup
from pypdf import PdfReader
from openpyxl import load_workbook

from app.config import (
    DOCS_DIR,
    CHUNK_SIZE,
    CHUNK_OVERLAP,
    EMBEDDING_MODEL,
    EMBEDDING_VERSION,
    MAX_FILE_SIZE_MB,
    MAX_CHUNKS,
)

SUPPORTED_EXTENSIONS = {".md", ".pdf", ".txt", ".html", ".xlsx"}
IGNORE_FILE_NAMES = {".env"}
IGNORE_DIRS = {".git", "node_modules", "__pycache__"}


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def normalize_text(text: str) -> str:
    lines = [line.strip() for line in text.splitlines()]
    lines = [line for line in lines if line]
    return "\n".join(lines)


def file_too_large(path: Path) -> bool:
    size_mb = path.stat().st_size / (1024 * 1024)
    return size_mb > MAX_FILE_SIZE_MB


def load_text_file(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")


def load_pdf(path: Path) -> str:
    reader = PdfReader(str(path))
    texts = []
    for page in reader.pages:
        texts.append(page.extract_text() or "")
    return "\n".join(texts)


def load_html(path: Path) -> str:
    html = path.read_text(encoding="utf-8", errors="ignore")
    soup = BeautifulSoup(html, "html.parser")
    return soup.get_text(separator="\n")


def load_xlsx(path: Path) -> str:
    wb = load_workbook(str(path), data_only=True)
    blocks = []
    for sheet in wb.worksheets:
        blocks.append(f"Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = [str(v) for v in row if v is not None]
            if values:
                blocks.append(" | ".join(values))
    return "\n".join(blocks)


def read_document(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".md", ".txt"}:
        return load_text_file(path)
    if suffix == ".pdf":
        return load_pdf(path)
    if suffix == ".html":
        return load_html(path)
    if suffix == ".xlsx":
        return load_xlsx(path)
    return ""


def chunk_text(text: str, chunk_size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> List[str]:
    chunks = []
    start = 0
    n = len(text)

    while start < n:
        end = min(n, start + chunk_size)
        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)
        if end >= n:
            break
        start = max(0, end - overlap)

        if len(chunks) >= MAX_CHUNKS:
            break

    return chunks


def discover_documents() -> List[Path]:
    documents: List[Path] = []

    for root, dirs, files in os.walk(DOCS_DIR):
        dirs[:] = [d for d in dirs if d not in IGNORE_DIRS]

        for file_name in files:
            if file_name in IGNORE_FILE_NAMES:
                continue

            path = Path(root) / file_name

            if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
                continue

            if file_too_large(path):
                continue

            documents.append(path)

    return documents


def build_chunks() -> List[Dict]:
    docs = discover_documents()
    all_chunks: List[Dict] = []

    for doc_index, path in enumerate(docs, start=1):
        raw_text = read_document(path)
        cleaned_text = normalize_text(raw_text)

        if not cleaned_text:
            continue

        doc_id = f"doc_{doc_index}"
        content_hash = sha256_text(cleaned_text)
        chunks = chunk_text(cleaned_text)

        for chunk_index, chunk in enumerate(chunks, start=1):
            all_chunks.append({
                "doc_id": doc_id,
                "chunk_id": f"{doc_id}_chunk_{chunk_index}",
                "file_path": str(path),
                "content_hash": content_hash,
                "embedding_model": EMBEDDING_MODEL,
                "embedding_version": EMBEDDING_VERSION,
                "text": chunk,
            })

    return all_chunks